"""
Streamlit App - Flatten, Gabung Akrual+SP2D & Visualisasi Laporan 16 Segmen
----------------------------------------------------------------------------
Upload file laporan Akrual (.xlsx) -> wajib.
Upload file laporan SP2D (.xlsx)   -> opsional, struktur segmen sama, hanya
                                       beda basis realisasi (kas vs akrual).

Kedua file digabung berdasarkan kode lengkap hierarki (kd_satker ... kd_item)
sehingga hasil akhirnya adalah satu tabel flat dengan kolom tambahan
'realisasi_sp2d' di sebelah 'realisasi_anggaran' (basis akrual).

Cara jalankan lokal (di PyCharm):
    1. pip install -r requirements.txt
    2. streamlit run app.py

Cara deploy publik (gratis):
    1. Push folder ini (app.py + requirements.txt) ke sebuah repo GitHub.
    2. Buka https://share.streamlit.io -> New app -> pilih repo & file app.py.
    3. Deploy.
"""

import io
import re

import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# Logika parsing (membaca hierarki dari laporan asli)
# ----------------------------------------------------------------------


def g(ws, r, c):
    return ws.cell(row=r, column=c).value


def parse_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    satker_kode = g(ws, 6, 15)
    satker_nama = g(ws, 6, 16)

    ctx = {
        "program_kode": None, "program_nama": None,
        "kegiatan_kode": None, "kegiatan_nama": None,
        "output_kode": None, "output_nama": None,          # level KRO
        "suboutput_kode": None, "suboutput_nama": None,    # level RO (Output)
        "komponen_kode": None, "komponen_nama": None,
        "subkomponen_kode": None, "subkomponen_nama": None,
        "akun_kode": None, "akun_nama": None,
    }

    rows_info = []

    for r in range(10, ws.max_row + 1):
        b = g(ws, r, 2)   # Program / Kegiatan
        c = g(ws, r, 3)   # KRO / RO(Output)
        e = g(ws, r, 5)   # Komponen
        f = g(ws, r, 6)   # SubKomponen
        h = g(ws, r, 8)   # Akun
        n = g(ws, r, 14)  # Item (kode + nama digabung)

        if b and str(b).startswith("*"):
            continue

        level = None
        if b is not None:
            level = "kegiatan" if "." in str(b) else "program"
        elif c is not None:
            level = "suboutput" if "." in str(c) else "output"
        elif e is not None:
            level = "komponen"
        elif f is not None:
            level = "subkomponen"
        elif h is not None:
            level = "akun"
        elif n is not None:
            level = "item"
        else:
            continue

        item_kode = None
        item_nama = None

        if level == "program":
            ctx["program_kode"] = b
            ctx["program_nama"] = g(ws, r, 4)
        elif level == "kegiatan":
            ctx["kegiatan_kode"] = b
            ctx["kegiatan_nama"] = g(ws, r, 9)
        elif level == "output":
            ctx["output_kode"] = c
            ctx["output_nama"] = g(ws, r, 7)
        elif level == "suboutput":
            ctx["suboutput_kode"] = c
            ctx["suboutput_nama"] = g(ws, r, 11)
        elif level == "komponen":
            ctx["komponen_kode"] = e
            ctx["komponen_nama"] = g(ws, r, 10)
        elif level == "subkomponen":
            ctx["subkomponen_kode"] = f
            ctx["subkomponen_nama"] = g(ws, r, 12)
        elif level == "akun":
            pagu_check = g(ws, r, 17)
            if (
                ctx["akun_kode"] == h
                and rows_info
                and rows_info[-1]["level"] == "akun"
                and rows_info[-1]["pagu"] == pagu_check
            ):
                ctx["akun_nama"] = (
                    str(ctx["akun_nama"]).rstrip() + " " + str(g(ws, r, 13)).strip()
                )
                rows_info[-1]["ctx"] = dict(ctx)
                continue
            ctx["akun_kode"] = h
            ctx["akun_nama"] = g(ws, r, 13)
        elif level == "item":
            item_full = n
            m = re.match(r"^\s*(\d+)\.\s*(.*)$", str(item_full))
            if m:
                item_kode, item_nama = m.group(1), m.group(2)
                is_continuation = False
            else:
                is_continuation = True
            if is_continuation and rows_info and rows_info[-1]["level"] == "item":
                rows_info[-1]["item_nama"] = (
                    str(rows_info[-1]["item_nama"]).rstrip() + " " + str(item_full).strip()
                )
                continue

        pagu = g(ws, r, 17)
        real_sd = g(ws, r, 26)

        rows_info.append({
            "row": r,
            "level": level,
            "ctx": dict(ctx),
            "item_kode": item_kode if level == "item" else None,
            "item_nama": item_nama if level == "item" else None,
            "pagu": pagu,
            "real_sd": real_sd,
        })

    leaf_rows = []
    n_total = len(rows_info)
    for i, ri in enumerate(rows_info):
        if ri["level"] == "item":
            leaf_rows.append(ri)
        elif ri["level"] == "akun":
            if i + 1 < n_total and rows_info[i + 1]["level"] == "item":
                continue
            leaf_rows.append(ri)

    header = {"satker_kode": satker_kode, "satker_nama": satker_nama}
    return leaf_rows, header


# ----------------------------------------------------------------------
# Definisi kolom output final (nama kolom pendek)
# ----------------------------------------------------------------------

OUTPUT_COLUMNS = [
    ("kd_satker", lambda ri, h: h["satker_kode"]),
    ("satker", lambda ri, h: h["satker_nama"]),
    ("kd_program", lambda ri, h: ri["ctx"]["program_kode"]),
    ("program", lambda ri, h: ri["ctx"]["program_nama"]),
    ("kd_keg", lambda ri, h: ri["ctx"]["kegiatan_kode"]),
    ("kegiatan", lambda ri, h: ri["ctx"]["kegiatan_nama"]),
    ("kd_kro", lambda ri, h: ri["ctx"]["output_kode"]),
    ("kro", lambda ri, h: ri["ctx"]["output_nama"]),
    ("kd_ro", lambda ri, h: ri["ctx"]["suboutput_kode"]),
    ("ro", lambda ri, h: ri["ctx"]["suboutput_nama"]),
    ("kd_komponen", lambda ri, h: ri["ctx"]["komponen_kode"]),
    ("komponen", lambda ri, h: ri["ctx"]["komponen_nama"]),
    ("kd_subkomponen", lambda ri, h: ri["ctx"]["subkomponen_kode"]),
    ("subkomponen", lambda ri, h: ri["ctx"]["subkomponen_nama"]),
    ("kd_akun", lambda ri, h: ri["ctx"]["akun_kode"]),
    ("akun", lambda ri, h: ri["ctx"]["akun_nama"]),
    ("kd_item", lambda ri, h: ri["item_kode"]),
    ("item", lambda ri, h: ri["item_nama"]),
    ("pagu_anggaran", lambda ri, h: ri["pagu"]),
    ("realisasi_anggaran", lambda ri, h: ri["real_sd"]),
]

# Kolom yang dipakai sebagai kunci gabung antara file Akrual & SP2D
# (mengidentifikasi Item secara unik lewat seluruh jenjang kodenya)
JOIN_KEY_COLS = [
    "kd_satker", "kd_program", "kd_keg", "kd_kro", "kd_ro",
    "kd_komponen", "kd_subkomponen", "kd_akun", "kd_item",
]

MONEY_COLS = {"pagu_anggaran", "realisasi_anggaran", "realisasi_sp2d"}

COLUMN_WIDTHS = {
    "kd_satker": 10, "satker": 30, "kd_program": 10, "program": 32,
    "kd_keg": 10, "kegiatan": 32, "kd_kro": 10, "kro": 30,
    "kd_ro": 10, "ro": 36, "kd_komponen": 10, "komponen": 32,
    "kd_subkomponen": 12, "subkomponen": 42, "kd_akun": 10, "akun": 28,
    "kd_item": 10, "item": 45, "pagu_anggaran": 16, "realisasi_anggaran": 18,
    "realisasi_sp2d": 16,
}


def leaf_rows_to_dataframe(leaf_rows, header):
    """Ubah leaf_rows jadi pandas DataFrame flat dengan nama kolom final."""
    records = []
    for ri in leaf_rows:
        row = {name: fn(ri, header) for name, fn in OUTPUT_COLUMNS}
        row["pagu_anggaran"] = row["pagu_anggaran"] or 0
        row["realisasi_anggaran"] = row["realisasi_anggaran"] or 0
        records.append(row)
    return pd.DataFrame(records, columns=[name for name, _ in OUTPUT_COLUMNS])


def merge_akrual_sp2d(df_akrual, df_sp2d):
    """Gabungkan df Akrual (basis utama) dengan df SP2D berdasarkan kode
    lengkap hierarki, menambahkan kolom 'realisasi_sp2d'.
    Mengembalikan (df_gabungan, info) di mana info berisi jumlah baris
    Item yang tidak ketemu pasangannya di masing-masing sisi."""
    sp2d_slim = df_sp2d[JOIN_KEY_COLS + ["realisasi_anggaran"]].rename(
        columns={"realisasi_anggaran": "realisasi_sp2d"}
    )
    # kalau ada duplikat kode Item yang sama persis di file SP2D, jumlahkan
    sp2d_slim = sp2d_slim.groupby(JOIN_KEY_COLS, dropna=False, as_index=False)["realisasi_sp2d"].sum()

    merged = df_akrual.merge(sp2d_slim, on=JOIN_KEY_COLS, how="left")
    unmatched_in_sp2d = int(merged["realisasi_sp2d"].isna().sum())
    merged["realisasi_sp2d"] = merged["realisasi_sp2d"].fillna(0)

    akrual_keys = set(map(tuple, df_akrual[JOIN_KEY_COLS].values))
    sp2d_keys = set(map(tuple, df_sp2d[JOIN_KEY_COLS].values))
    only_in_sp2d = len(sp2d_keys - akrual_keys)

    info = {"unmatched_in_sp2d": unmatched_in_sp2d, "only_in_sp2d": only_in_sp2d}
    return merged, info


def build_output_bytes(df, columns_order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Flat"

    font_name = "Arial"
    header_font = Font(name=font_name, bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell_font = Font(name=font_name, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, name in enumerate(columns_order, start=1):
        cell = ws.cell(row=1, column=j, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    money_col_idx = [j for j, name in enumerate(columns_order, start=1) if name in MONEY_COLS]

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, name in enumerate(columns_order, start=1):
            cell = ws.cell(row=i, column=j, value=row[name])
            cell.font = cell_font
            cell.border = border
            if j in money_col_idx:
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns_order))}{len(df) + 1}"

    for j, name in enumerate(columns_order, start=1):
        ws.column_dimensions[get_column_letter(j)].width = COLUMN_WIDTHS.get(name, 14)

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ----------------------------------------------------------------------
# Helper visualisasi
# ----------------------------------------------------------------------

def grouped_bar(df, group_col, value_cols, value_labels, top_n=None):
    agg = df.groupby(group_col, dropna=False)[value_cols].sum().reset_index()
    agg[group_col] = agg[group_col].fillna("(Tidak diketahui)")
    sort_col = value_cols[-1]
    agg = agg.sort_values(sort_col, ascending=False)

    if top_n and len(agg) > top_n:
        head = agg.iloc[:top_n]
        rest = agg.iloc[top_n:][value_cols].sum()
        rest_row = pd.DataFrame([{group_col: f"Lainnya ({len(agg) - top_n} item)", **rest.to_dict()}])
        agg = pd.concat([head, rest_row], ignore_index=True)

    melted = agg.melt(id_vars=[group_col], value_vars=value_cols, var_name="Jenis", value_name="Nilai")
    melted["Jenis"] = melted["Jenis"].map(value_labels)

    color_map = {"Pagu Anggaran": "#94A3B8", "Realisasi Anggaran (Akrual)": "#1F4E78", "Realisasi SP2D": "#2E9E5B"}

    fig = px.bar(
        melted, x="Nilai", y=group_col, color="Jenis", orientation="h",
        barmode="group", text_auto=".2s", color_discrete_map=color_map,
    )
    fig.update_layout(
        yaxis={"categoryorder": "total ascending"},
        legend_title_text="", xaxis_title="Nilai (Rp)", yaxis_title="",
        height=max(320, 40 * len(agg)),
        margin=dict(l=10, r=10, t=30, b=10),
    )
    return fig, agg


# ----------------------------------------------------------------------
# UI Streamlit
# ----------------------------------------------------------------------

st.set_page_config(page_title="Flatten & Visualisasi Laporan 16 Segmen", page_icon="📊", layout="wide")

st.title("📊 Flatten & Visualisasi Laporan Fa Detail (16 Segmen)")
st.write(
    "Upload laporan **Akrual** (wajib) dan opsional laporan **SP2D** dengan struktur segmen "
    "yang sama. Kedua file akan digabung menjadi satu tabel flat, ditambah kolom "
    "`realisasi_sp2d` bila file SP2D diisi."
)

col_up1, col_up2 = st.columns(2)
with col_up1:
    file_akrual = st.file_uploader(
        "1️⃣ Upload File Akrual (wajib)", type=["xlsx"], key="akrual",
        help='Contoh nama file: "Laporan Fa Detail (16 Segmen)-horti-akrual-060826.xlsx"',
    )
with col_up2:
    file_sp2d = st.file_uploader(
        "2️⃣ Upload File SP2D (opsional)", type=["xlsx"], key="sp2d",
        help='Contoh nama file: "Laporan Fa Detail (16 Segmen)-horti-sp2d-060826.xlsx"',
    )

if file_akrual is not None:
    try:
        with st.spinner("Memproses file Akrual..."):
            leaf_rows_a, header_a = parse_workbook(io.BytesIO(file_akrual.getvalue()))
            df_akrual = leaf_rows_to_dataframe(leaf_rows_a, header_a)

        df_final = df_akrual
        columns_order = [name for name, _ in OUTPUT_COLUMNS]
        info = None

        if file_sp2d is not None:
            with st.spinner("Memproses file SP2D & menggabungkan..."):
                leaf_rows_s, header_s = parse_workbook(io.BytesIO(file_sp2d.getvalue()))
                df_sp2d = leaf_rows_to_dataframe(leaf_rows_s, header_s)
                df_final, info = merge_akrual_sp2d(df_akrual, df_sp2d)
                columns_order = columns_order + ["realisasi_sp2d"]

        st.success(f"Berhasil! {len(df_final)} baris data flat dihasilkan.")

        if info is not None:
            if info["unmatched_in_sp2d"] > 0:
                st.warning(
                    f"⚠️ {info['unmatched_in_sp2d']} baris Item di file Akrual tidak ditemukan "
                    "pasangannya di file SP2D (realisasi_sp2d diisi 0 untuk baris tsb)."
                )
            if info["only_in_sp2d"] > 0:
                st.warning(
                    f"⚠️ {info['only_in_sp2d']} baris Item ada di file SP2D tapi tidak ada di file "
                    "Akrual, sehingga tidak ikut ditampilkan (tabel mengikuti baris file Akrual)."
                )
            if info["unmatched_in_sp2d"] == 0 and info["only_in_sp2d"] == 0:
                st.info("✅ Seluruh Item cocok sempurna antara file Akrual dan SP2D.")

        total_pagu = df_final["pagu_anggaran"].sum()
        total_real = df_final["realisasi_anggaran"].sum()
        total_sisa = total_pagu - total_real
        pct_overall = (total_real / total_pagu * 100) if total_pagu else 0

        if info is not None:
            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Total Pagu Anggaran", f"Rp {total_pagu:,.0f}")
            col2.metric("Realisasi (Akrual)", f"Rp {total_real:,.0f}")
            col3.metric("Realisasi (SP2D)", f"Rp {df_final['realisasi_sp2d'].sum():,.0f}")
            col4.metric("Sisa Anggaran", f"Rp {total_sisa:,.0f}")
            col5.metric("% Realisasi Akrual", f"{pct_overall:.2f}%")
        else:
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Pagu Anggaran", f"Rp {total_pagu:,.0f}")
            col2.metric("Total Realisasi Anggaran", f"Rp {total_real:,.0f}")
            col3.metric("Sisa Anggaran", f"Rp {total_sisa:,.0f}")
            col4.metric("% Realisasi", f"{pct_overall:.2f}%")

        st.progress(min(pct_overall / 100, 1.0))
        st.caption(
            "Cocokkan Total Pagu & Realisasi di atas dengan baris 'JUMLAH SELURUHNYA' pada "
            "laporan asli untuk memastikan tidak ada data yang hilang atau terhitung dobel."
        )

        base_name = file_akrual.name.rsplit(".", 1)[0].replace("-akrual", "").replace("akrual", "")
        output_buffer = build_output_bytes(df_final, columns_order)
        csv_bytes = df_final[columns_order].to_csv(index=False, sep=";").encode("utf-8-sig")

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                "⬇️ Unduh file hasil (Excel .xlsx)", data=output_buffer,
                file_name=f"{base_name}_FLAT.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_dl2:
            st.download_button(
                "⬇️ Unduh file hasil (CSV .csv)", data=csv_bytes,
                file_name=f"{base_name}_FLAT.csv", mime="text/csv",
                use_container_width=True,
            )
        st.caption(
            "File CSV menggunakan pemisah titik-koma (;) agar kolom tidak pecah saat dibuka "
            "langsung di Excel versi Indonesia."
        )

        st.divider()
        st.subheader("📈 Visualisasi Realisasi Anggaran")

        if info is not None:
            value_cols = ["pagu_anggaran", "realisasi_anggaran", "realisasi_sp2d"]
            value_labels = {
                "pagu_anggaran": "Pagu Anggaran",
                "realisasi_anggaran": "Realisasi Anggaran (Akrual)",
                "realisasi_sp2d": "Realisasi SP2D",
            }
        else:
            value_cols = ["pagu_anggaran", "realisasi_anggaran"]
            value_labels = {"pagu_anggaran": "Pagu Anggaran", "realisasi_anggaran": "Realisasi Anggaran (Akrual)"}

        tab_keg, tab_ro, tab_komp, tab_akun = st.tabs(
            ["Per Kegiatan", "Per Output (RO)", "Per Komponen", "Per Akun"]
        )
        with tab_keg:
            fig, agg = grouped_bar(df_final, "kegiatan", value_cols, value_labels)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("Lihat tabel"):
                st.dataframe(agg, use_container_width=True, hide_index=True)
        with tab_ro:
            fig, agg = grouped_bar(df_final, "ro", value_cols, value_labels)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("Lihat tabel"):
                st.dataframe(agg, use_container_width=True, hide_index=True)
        with tab_komp:
            fig, agg = grouped_bar(df_final, "komponen", value_cols, value_labels, top_n=15)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("Lihat tabel"):
                st.dataframe(agg, use_container_width=True, hide_index=True)
        with tab_akun:
            fig, agg = grouped_bar(df_final, "akun", value_cols, value_labels, top_n=15)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("Lihat tabel"):
                st.dataframe(agg, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("🔍 Pratinjau Data Flat")
        st.dataframe(df_final[columns_order].head(50), use_container_width=True, hide_index=True)

    except Exception as exc:
        st.error(
            "Terjadi kesalahan saat memproses file. Pastikan format file sesuai dengan "
            "'Laporan Fa Detail (16 Segmen)' dari SAKTI/SPAN."
        )
        st.exception(exc)
else:
    st.info("Silakan upload minimal file Akrual untuk mulai memproses.")
